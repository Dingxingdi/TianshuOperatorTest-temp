#!/usr/bin/env python3
"""
为第一批及格算子国产GPU测试.xlsx 中失败的算子添加"失败原因"列。
读取 test_results_20260509_141455/ 目录下的日志文件，解析失败原因。

使用方法：
    cd /root/JudeWorkplace/TianshuOperatorTest
    python3 add_failure_reasons.py
"""

import os
import re
import openpyxl

# 路径配置
EXCEL_PATH = "TianshuOperatorTest/第一批及格算子国产GPU测试.xlsx"
LOG_DIR = "TianshuOperatorTest/test_results_20260509_141455/"
NEW_COLUMN_HEADER = "失败原因"

# 日志目录下的所有文件名
TEST_LOG_FILES = [f for f in os.listdir(LOG_DIR) if f.endswith("_test.log")]


def build_log_map():
    """
    构建算子名到日志文件的映射。
    算子名可能格式: aten::xxx, xxx_（结尾有下划线）, 友好名称
    日志文件名格式: xxx_test.log
    """
    log_map = {}
    for fname in TEST_LOG_FILES:
        base = fname[:-len("_test.log")]
        log_map[base] = os.path.join(LOG_DIR, fname)
        log_map[base.lower()] = os.path.join(LOG_DIR, fname)
        # 处理尾部下划线
        base_no_underscore = base.rstrip('_')
        if base_no_underscore != base:
            log_map[base_no_underscore] = os.path.join(LOG_DIR, fname)
            log_map[base_no_underscore.lower()] = os.path.join(LOG_DIR, fname)
    return log_map


def normalize_operator_name(name):
    """标准化算子名称，以便匹配日志文件。"""
    if name is None:
        return None
    name = str(name).strip()
    if name.startswith("aten::"):
        name = name[6:]
    return name


def find_log_file(op_name, log_map):
    """尝试多种匹配方式找到对应的日志文件"""
    if op_name is None:
        return None
    
    # 直接匹配
    if op_name in log_map:
        return log_map[op_name]
    
    # 小写匹配
    if op_name.lower() in log_map:
        return log_map[op_name.lower()]
    
    # 去掉尾部下划线
    name_clean = op_name.rstrip('_')
    if name_clean != op_name:
        if name_clean in log_map:
            return log_map[name_clean]
        if name_clean.lower() in log_map:
            return log_map[name_clean.lower()]
    
    # 去掉前导下划线
    name_no_prefix = op_name.lstrip('_')
    if name_no_prefix != op_name:
        if name_no_prefix in log_map:
            return log_map[name_no_prefix]
        if name_no_prefix.lower() in log_map:
            return log_map[name_no_prefix.lower()]
    
    return None


def parse_status(content):
    """提取日志中的状态信息"""
    # 检查 TIMEOUT
    timeout_match = re.search(r'# Status: TIMEOUT.*?(\d+)s', content)
    if timeout_match:
        return "TIMEOUT", int(timeout_match.group(1))
    
    # 检查 Exit Code
    exit_match = re.search(r'# Exit Code: (-?\d+)', content)
    if exit_match:
        return "EXIT", int(exit_match.group(1))
    
    return "UNKNOWN", None


def parse_test_counts(content):
    """获取 PASSED / FAILED 数量"""
    passes = len(re.findall(r' PASSED ', content))
    fails = len(re.findall(r' FAILED ', content))
    return passes, fails


def read_log_content(log_path):
    """智能读取日志文件：先读开头500KB获取header/meta，若未失败则再读末尾500KB获取错误信息"""
    header = ""
    tail = ""
    try:
        with open(log_path, 'r', encoding='utf-8', errors='replace') as f:
            header = f.read(500000)  # 读前500KB
    except Exception as e:
        pass
    
    try:
        with open(log_path, 'rb') as f:
            f.seek(0, 2)  # 跳到文件末尾
            filesize = f.tell()
            read_size = min(filesize, 600000)  # 读末尾600KB
            f.seek(filesize - read_size)
            tail = f.read(read_size).decode('utf-8', errors='replace')
    except Exception as e:
        pass
    
    return header, tail


def parse_failure_reason(log_path, bench_log_path=None):
    """解析日志文件，提取失败原因"""
    header, tail = read_log_content(log_path)
    content = header + "\n" + tail  # 合并首尾，优先用尾部
    
    if not content.strip():
        return "读取日志失败: 文件为空"
    
    # 1. 检查状态 (TIMEOUT / Exit Code)
    status, status_val = parse_status(header)  # header通常包含meta信息
    
    if status == "TIMEOUT":
        return f"测试超时({status_val}秒)，可能算子卡死或Triton编译无限循环"
    
    # 2. 计算 PASSED/FAILED
    passes = len(re.findall(r' PASSED ', content))
    fails = len(re.findall(r' FAILED ', content))
    
    if status_val == 0 and fails == 0:
        # Exit code 0, no failures -> 检查benchmark是否有问题
        if bench_log_path and os.path.exists(bench_log_path):
            try:
                with open(bench_log_path, 'r', encoding='utf-8', errors='replace') as bf:
                    bench_content = bf.read(5000)
                bench_exit = re.search(r'# Exit Code: (\d+)', bench_content)
                if bench_exit and int(bench_exit.group(1)) != 0:
                    return f"准确率测试通过但benchmark未成功运行(Exit Code {bench_exit.group(1)})——Benchmark脚本执行异常，结果不可信"
            except:
                pass
        return "测试通过但被标记为失败(数据可能不准确)"
    
    if status_val == -6:
        return "进程被信号终止(SIGABRT/SIGSEGV)，算子执行时崩溃"
    
    if status_val == 5:
        return "测试执行错误(Exit Code 5)，可能无匹配测试用例或pytest配置问题"
    
    # 3. 检查各种错误模式
    
    # 3a. cuDNN/IXDNN 错误 (最高优先级，通常是根本原因)
    cudnn_match = re.search(r'RuntimeError: cuDNN error: (.*?)[\n\r]', content)
    if cudnn_match:
        err_detail = cudnn_match.group(1).strip()
        if 'IXDNN_STATUS_BAD_PARAM' in err_detail:
            return "cuDNN参数错误(IXDNN_STATUS_BAD_PARAM)——天数GPU cuDNN库不支持该算子参数"
        elif 'IXDNN_STATUS' in err_detail:
            return f"IXDNN错误({err_detail})"
        else:
            return "cuDNN运行时错误"
    
    # 3b. NotImplementedError
    not_impl_match = re.search(r'NotImplementedError:.*?[\n\r]', content)
    if not_impl_match:
        return "该算子在天数GPU上无CUDA实现"
    
    # 3c. cuBLAS 非RuntimeError形式的错误
    cublas_stderr = re.search(r'cuBLAS error at.*?(?:\n|$)', content)
    if cublas_stderr:
        return f"运行时错误: cuBLAS内部错误({cublas_stderr.group(0).strip()[:100]})"
    
    # 3d. Triton 编译错误 (SIToFP vector length mismatch等)
    triton_compile_err = re.search(r'SIToFP source and dest vector length mismatch', content)
    if triton_compile_err:
        return "运行时错误: Triton LLVM IR编译错误(SIToFP向量长度不匹配)——Triton在天数GPU上LLVM IR生成有问题"
    
    # 3e. RuntimeError 提取详细信息
    runtime_errors = []
    for m in re.finditer(r"(?:RuntimeError|Error): ([^\n\r]+)", content):
        err_msg = m.group(1).strip()[:120]
        if err_msg not in runtime_errors:
            runtime_errors.append(err_msg)
    
    # 分类处理 RuntimeError
    for err in runtime_errors:
        if "gemm of double is not supported" in err:
            return "运行时错误: 天数CoreX不支持double类型的gemm操作(CoreX对double精度支持有限)"
        if "CUBLAS_STATUS_NOT_SUPPORTED" in err:
            short_err = err[:80]
            return f"运行时错误: cuBLAS不支持该操作({short_err})"
        if "ixFFT doesn't support" in err:
            return f"运行时错误: 天数ixFFT不支持该数据类型"
        if "PassManager::run failed" in err:
            return "运行时错误: Triton PassManager编译失败"
        if "not implemented for" in err:
            return f"运行时错误: {err}"
    
    # 3f. AssertionError (精度不匹配)
    assertion_matches = re.findall(
        r'Mismatched elements: (\d+) / (\d+) \((\d+\.?\d*)%\)',
        content
    )
    if assertion_matches:
        max_pct = max(float(m[2]) for m in assertion_matches)
        # 找到百分比最大的那一条，提取具体元素数
        worst_match = max(assertion_matches, key=lambda m: float(m[2]))
        mismatched = int(worst_match[0])
        total = int(worst_match[1])
        
        if max_pct > 99:
            return f"精度不匹配(最高{max_pct:.1f}%元素不匹配)——FlagGems Triton kernel计算结果与标准PyTorch几乎完全不同"
        elif max_pct > 50:
            return f"精度不匹配(最高{max_pct:.1f}%元素不匹配)——FlagGems实现存在较大精度偏差"
        elif max_pct > 0.05:
            return f"精度不匹配(最高{max_pct:.2f}%元素不匹配)——FlagGems实现存在轻微精度偏差"
        else:
            # 百分比 < 0.05% 时，显示实际元素数更清晰
            real_pct = (mismatched / total * 100) if total > 0 else 0
            return f"精度不匹配({mismatched}/{total}元素不匹配, {real_pct:.4f}%)——FlagGems实现存在极微小精度偏差"
    
    # Check for "Scalars are not close" (without Mismatched elements line)
    if fails > 0 and re.search(r"Scalars are not close", content):
        return "运行时错误: Scalars are not close!"
    
    # 3g. 全部FAILED (90/90等)
    if fails > 0 and passes == 0:
        return f"测试全部失败[{fails}失败/0通过]，FlagGems实现存在严重bug"
    
    # 3e. 有FAILED但没有AssertionError详细信息被提取到
    if fails > 0:
        # 尝试找short test summary
        summary_match = re.search(r'=============+ short test summary info =============+\n(.*?)(?:\n=+|$)', content, re.DOTALL)
        if summary_match:
            failed_tests = summary_match.group(1).strip()
            lines = [l.strip() for l in failed_tests.split('\n') if l.strip()]
            combined_err = "\n".join(lines[:5])
            return f"测试失败，详见日志[{fails}失败/{passes + fails}总]: {combined_err[:200]}"
        
        return f"测试失败[{fails}失败/{passes + fails}总]，FlagGems实现存在bug"
    
    # 3f. 保底
    if runtime_errors:
        return f"运行时错误: {runtime_errors[0][:100]}"
    
    return "未知错误(请查看原始日志细节)"



def get_all_failed_ops(ws):
    """获取sheet中所有失败的算子"""
    failed_ops = []
    for row_idx in range(2, ws.max_row + 1):
        op_name = ws.cell(row=row_idx, column=1).value
        result = ws.cell(row=row_idx, column=3).value
        if op_name and result and isinstance(result, str) and '失败' in result:
            failed_ops.append((row_idx, str(op_name).strip(), str(result).strip()))
    return failed_ops


def main():
    print("=" * 60)
    print("开始分析失败的算子...")
    print("=" * 60)
    
    print(f"\n加载Excel文件: {EXCEL_PATH}")
    wb = openpyxl.load_workbook(EXCEL_PATH)
    
    log_map = build_log_map()
    print(f"日志目录: {LOG_DIR}")
    print(f"找到 {len(TEST_LOG_FILES)} 个测试日志文件")
    
    total_failed = 0
    total_found_log = 0
    
    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        print(f"\n--- Sheet: {sheet_name} ---")
        
        # 检查是否已有"失败原因"列
        existing_headers = [ws.cell(row=1, column=c).value for c in range(1, ws.max_column + 1)]
        
        if NEW_COLUMN_HEADER in existing_headers:
            new_col_idx = existing_headers.index(NEW_COLUMN_HEADER) + 1
            print(f"  已有'{NEW_COLUMN_HEADER}'列 (列{new_col_idx})")
        else:
            new_col_idx = ws.max_column + 1
            ws.cell(row=1, column=new_col_idx, value=NEW_COLUMN_HEADER)
            print(f"  新增'{NEW_COLUMN_HEADER}'列 (列{new_col_idx})")
        
        failed_ops = get_all_failed_ops(ws)
        print(f"  失败算子数: {len(failed_ops)}")
        
        sheet_found = 0
        for row_idx, op_name, result_val in failed_ops:
            total_failed += 1
            
            norm_name = normalize_operator_name(op_name)
            log_path = find_log_file(norm_name, log_map)
            
            # 如果没找到，尝试直接文件名
            if not log_path:
                alt_path = os.path.join(LOG_DIR, f"{norm_name}_test.log")
                if os.path.exists(alt_path):
                    log_path = alt_path
                else:
                    # 模糊匹配
                    for fname in TEST_LOG_FILES:
                        fbase = fname[:-len("_test.log")]
                        if norm_name.replace("_", "").lower() == fbase.replace("_", "").lower():
                            log_path = os.path.join(LOG_DIR, fname)
                            break
                    if not log_path:
                        for fname in TEST_LOG_FILES:
                            fbase = fname[:-len("_test.log")]
                            if norm_name.lower() in fbase.lower() or fbase.lower() in norm_name.lower():
                                log_path = os.path.join(LOG_DIR, fname)
                                break
            
            if log_path:
                sheet_found += 1
                total_found_log += 1
                # 查找对应的 benchmark 日志文件
                bench_log_path = None
                base_name = os.path.splitext(os.path.basename(log_path))[0].replace("_test", "_bench")
                bench_file = os.path.join(LOG_DIR, base_name + ".log")
                if os.path.exists(bench_file):
                    bench_log_path = bench_file
                reason = parse_failure_reason(log_path, bench_log_path)
            else:
                reason = "未找到测试日志文件"
            
            ws.cell(row=row_idx, column=new_col_idx, value=reason)
            print(f"  [{row_idx}] {op_name:45s} → {reason}")
        
        print(f"  Sheet内找到日志: {sheet_found}/{len(failed_ops)}")
    
    print(f"\n{'=' * 60}")
    print(f"总计失败算子: {total_failed}")
    print(f"找到日志文件: {total_found_log}")
    print(f"未找到日志: {total_failed - total_found_log}")
    print(f"\n保存到: {EXCEL_PATH}")
    wb.save(EXCEL_PATH)
    print("完成!")


if __name__ == "__main__":
    main()
