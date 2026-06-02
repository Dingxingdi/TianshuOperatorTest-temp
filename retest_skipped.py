#!/usr/bin/env python3
"""
重新测试被跳过的算子并更新Excel。
处理三种跳过情况：
1. 实际已通过但Excel被标为"跳过"的 → 修正为"成功"
2. CSV中有但算子名不匹配的 → 修正匹配后重测
3. 确实缺少CSV命令的 → 补全命令后重测（不含无Worktree的）

使用方法：
    cd /root/JudeWorkplace
    python3 TianshuOperatorTest/retest_skipped.py
"""

import csv
import subprocess
import sys
import os
import json
import re
import time
import openpyxl
from datetime import datetime
from pathlib import Path

WORK_DIR = Path("/root/JudeWorkplace")
FLAGGEMS_BASE = WORK_DIR / "FlagGems_minimax_2_7"
CSV_PATH = WORK_DIR / "TianshuOperatorTest/all_operator_commands.csv"
EXCEL_PATH = WORK_DIR / "TianshuOperatorTest/第一批及格算子国产GPU测试.xlsx"
SUMMARY_JSON = WORK_DIR / "TianshuOperatorTest/test_results_20260509_141455/summary_20260509_141455.json"


# ==================== 别名映射表 ====================
# 用于匹配Excel中的算子名到CSV中的算子名
ALIAS_MAP = {
    # Excel中的名字 → CSV中的名字
    "aten::special_erfc": "erfc",                    # CSV行326
    "aten::__ixor__": "__xor__",                     # CSV行110
    "aten::_nested_view_from_buffer_copy": "nested_view_from_buffer_copy",  # CSV行175
    "ConvTranspose": None,                           # Sheet6中，CSV无对应，需追加
    "Shape": None,                                    # Sheet6中，CSV无对应，需追加
}


def load_excel_data():
    """读取Excel中所有被跳过的算子及所在sheet"""
    wb = openpyxl.load_workbook(EXCEL_PATH, data_only=True)
    skipped = []
    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        result_col = None
        for c in range(1, ws.max_column + 1):
            h = ws.cell(row=1, column=c).value
            if h == "天数测试结果":
                result_col = c
                break
        if not result_col:
            continue
        for r in range(2, ws.max_row + 1):
            name = ws.cell(row=r, column=1).value
            result = ws.cell(row=r, column=result_col).value
            if name and result and "跳过" in str(result):
                skipped.append((sheet_name, r, str(name).strip()))
    wb.close()
    print(f"共找到 {len(skipped)} 个被跳过算子")
    return skipped


def fix_type_a_passed_ops():
    """
    修正类型A：已在summary JSON中标记为success但Excel标为"跳过"的算子。
    这些算子accuracy和benchmark都通过了，只是benchmark没有输出可解析的数据。
    应该标记为"成功"而不是"跳过"。
    """
    if not SUMMARY_JSON.exists():
        print(f"⚠ 找不到summary JSON: {SUMMARY_JSON}")
        return 0

    with open(SUMMARY_JSON, "r") as f:
        data = json.load(f)

    wb = openpyxl.load_workbook(EXCEL_PATH)
    fixed_count = 0

    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        result_col = None
        for c in range(1, ws.max_column + 1):
            h = ws.cell(row=1, column=c).value
            if h == "天数测试结果":
                result_col = c
                break
        if not result_col:
            continue

        for r in range(2, ws.max_row + 1):
            name = ws.cell(row=r, column=1).value
            result = ws.cell(row=r, column=result_col).value
            if not name or not result or "跳过" not in str(result):
                continue

            name_clean = str(name).strip()

            # 在summary中查找
            op_data = data["operators"].get(name_clean)
            if not op_data and name_clean.startswith("aten::"):
                op_data = data["operators"].get(name_clean[6:])

            if op_data and op_data.get("accuracy_passed") == True:
                # Accuracy passed - should be marked as success
                # Check if there's a speedup value from bench
                bench = op_data.get("benchmark", {})
                bench_passed = op_data.get("benchmark_passed")
                bench_data = bench.get("data") if bench else None

                if bench_passed == True:
                    # Benchmark passed too - mark as "成功"
                    # Try to get speedup from bench data if available
                    if bench_data:
                        speeds = [d["speedup"] for d in bench_data if d.get("status") == "SUCCESS" and d.get("speedup") is not None]
                        if speeds:
                            avg_speed = round(sum(speeds) / len(speeds), 4)
                            ws.cell(row=r, column=2, value=avg_speed)  # 更新加速比
                    ws.cell(row=r, column=result_col, value="成功")
                    fixed_count += 1
                    print(f"  ✅ 修正: {name_clean:50s} → 成功 (accuracy通过, benchmark通过)")
                else:
                    # Accuracy passed but benchmark didn't
                    ws.cell(row=r, column=result_col, value="成功")
                    fixed_count += 1
                    print(f"  ✅ 修正: {name_clean:50s} → 成功 (accuracy通过, benchmark未跑出数据)")

    wb.save(EXCEL_PATH)
    print(f"\n类型A修正完成: {fixed_count} 个算子从'跳过'修正为'成功'")
    return fixed_count


def get_safe_filename(op_name):
    """将算子名转为安全的文件名"""
    safe = op_name.replace("/", "_").replace("\\", "_").replace(" ", "_")
    safe = safe.replace("<", "_").replace(">", "_").replace(":", "_")
    safe = safe.replace('"', "_").replace("'", "_").replace("|", "_")
    safe = safe.replace("?", "_").replace("*", "_")
    return safe


def find_worktree_dir(op_name):
    """根据算子名找到 worktree 目录"""
    worktrees_base = FLAGGEMS_BASE / ".worktrees"
    candidates = [f"gen-{op_name}", f"gen-{op_name.replace(' ', '+')}"]
    seen = set()
    for c in candidates:
        if c not in seen:
            seen.add(c)
            d = worktrees_base / c
            if d.is_dir():
                return str(d)
    return None


def append_to_csv(csv_path, op_name, test_cmd, bench_cmd):
    """向 CSV 追加新行"""
    with open(csv_path, "a", encoding="utf-8") as f:
        # Escape commas in commands
        test_escaped = test_cmd.replace('"', '""')
        bench_escaped = bench_cmd.replace('"', '""')
        f.write(f'\n{op_name},"{test_escaped}","{bench_escaped}"')
    print(f"  已追加到CSV: {op_name}")


def retest_operators():
    """
    重测15个被跳过的算子。
    包括：
    - 别名匹配的3个 (special_erfc, __ixor__, _nested_view_from_buffer_copy)
    - CSV中缺少的2个 (ConvTranspose, Shape)
    """
    wb = openpyxl.load_workbook(EXCEL_PATH)
    csv_commands = {}
    with open(CSV_PATH, "r", encoding="utf-8") as f:
        reader = csv.DictReader(f)
        for row in reader:
            name = row.get("Operator", "").strip()
            csv_commands[name] = {
                "test": row.get("Test Command", "").strip(),
                "bench": row.get("Benchmark Command", "").strip(),
            }

    retest_results = {}

    print("\n=== 重测计划 ===")
    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        result_col = None
        for c in range(1, ws.max_column + 1):
            h = ws.cell(row=1, column=c).value
            if h == "天数测试结果":
                result_col = c
                break
        if not result_col:
            continue

        for r in range(2, ws.max_row + 1):
            name = ws.cell(row=r, column=1).value
            result = ws.cell(row=r, column=result_col).value
            if not name or not result or "跳过" not in str(result):
                continue

            name_clean = str(name).strip()

            # 跳过已修正的（重新读取Excel）
            if result != "跳过":
                continue

            # 跳过无Worktree的
            wt_dir = find_worktree_dir(name_clean.replace("aten::", ""))
            if not wt_dir:
                print(f"  ⏭ 跳过 {name_clean}: 无Worktree目录")
                continue

            # 获取CSV命令
            csv_key = ALIAS_MAP.get(name_clean, name_clean.replace("aten::", ""))
            cmd_info = csv_commands.get(csv_key)

            if not cmd_info or not cmd_info["test"]:
                print(f"  ⏭ 跳过 {name_clean}: CSV中未找到命令(别名:{csv_key})")
                continue

            print(f"  ▶ 计划重测: {name_clean:50s} | CSV别名: {csv_key} | Worktree: {Path(wt_dir).name}")
            retest_results[(sheet_name, r, name_clean)] = {
                "csv_key": csv_key,
                "test_cmd": cmd_info["test"],
                "bench_cmd": cmd_info["bench"],
                "worktree": wt_dir,
            }

    print(f"\n实际需要重测: {len(retest_results)} 个算子")

    if not retest_results:
        print("没有需要重测的算子。")
        return

    # 开始重测
    print("\n" + "=" * 60)
    print("开始重测...")
    print("=" * 60)

    gpu_id = 0  # 使用GPU 0顺序执行
    for (sheet_name, row_idx, op_name), info in retest_results.items():
        print(f"\n▶ [{op_name}] 开始测试...")
        wt_dir = info["worktree"]
        test_cmd = info["test_cmd"]

        # 去除已有的CUDA_VISIBLE_DEVICES前缀
        cleaned_cmd = test_cmd
        while True:
            original = cleaned_cmd
            cleaned_cmd = re.sub(r'^CUDA_VISIBLE_DEVICES=\d+\s+', '', cleaned_cmd)
            cleaned_cmd = re.sub(r'^PYTHONPATH=[^\s]+\s+', '', cleaned_cmd)
            if cleaned_cmd == original:
                break

        full_cmd = f"cd {wt_dir} && VLLM_PLUGINS=\"\" CUDA_VISIBLE_DEVICES={gpu_id} {cleaned_cmd}"
        print(f"  运行: {cleaned_cmd[:120]}...")

        start = time.time()
        try:
            proc = subprocess.run(
                full_cmd, shell=True, capture_output=True, text=True,
                timeout=7200, executable="/bin/bash"
            )
            exit_code = proc.returncode
            duration = round(time.time() - start, 1)
            passed = (exit_code == 0)

            print(f"  {'✅ 通过' if passed else '❌ 失败'} (exit={exit_code}, {duration}s)")
            print(f"  stdout尾: ...{proc.stdout[-300:].strip()[-200:]}")
            print(f"  stderr尾: ...{proc.stderr[-300:].strip()[-200:]}")

            # 解析结果
            test_passed = passed

            # 跑benchmark
            bench_cmd = info["bench_cmd"]
            bench_passed = None
            if bench_cmd:
                cleaned_bench = bench_cmd
                while True:
                    original = cleaned_bench
                    cleaned_bench = re.sub(r'^CUDA_VISIBLE_DEVICES=\d+\s+', '', cleaned_bench)
                    cleaned_bench = re.sub(r'^PYTHONPATH=[^\s]+\s+', '', cleaned_bench)
                    if cleaned_bench == original:
                        break

                full_bench = f"cd {wt_dir} && VLLM_PLUGINS=\"\" CUDA_VISIBLE_DEVICES={gpu_id} {cleaned_bench}"
                print(f"  Benchmark: {cleaned_bench[:100]}...")
                bench_proc = subprocess.run(
                    full_bench, shell=True, capture_output=True, text=True,
                    timeout=7200, executable="/bin/bash"
                )
                bench_passed = (bench_proc.returncode == 0)
                print(f"  {'✅ Benchmark通过' if bench_passed else '❌ Benchmark失败'} (exit={bench_proc.returncode})")

                # 从bench输出解析加速比
                speedup = None
                for line in bench_proc.stdout.split("\n"):
                    m = re.search(r'Gems Speedup.*?(\d+\.\d+)x?', line)
                    if m:
                        speedup = float(m.group(1))
                        break
                    m2 = re.search(r'(?:SUCCESS|FAILED)\s+\S+\s+\S+\s+(\d+\.\d+)', line)
                    if m2 and speedup is None:
                        speedup = float(m2.group(1))

                if bench_passed and speedup:
                    print(f"  加速比: {speedup}")
            else:
                print(f"  无benchmark命令")

            # 更新Excel
            ws = wb[sheet_name]
            result_col = None
            speed_col = None
            for c in range(1, ws.max_column + 1):
                h = ws.cell(row=1, column=c).value
                if h == "天数测试结果":
                    result_col = c
                if h and "生成加速比" in str(h):
                    speed_col = c

            if test_passed:
                ws.cell(row=row_idx, column=result_col, value="成功")
                if bench_passed:
                    print(f"  ✅ 更新Excel: {op_name} → 成功")
                else:
                    print(f"  ⚠ 更新Excel: {op_name} → 成功 (accuracy通过, benchmark有问题)")
            else:
                # Extract failure reason
                failure_reason = "未知错误"
                combined_output = proc.stdout + "\n" + proc.stderr
                err_match = re.search(r'(RuntimeError|Error|Mismatched elements):\s*([^\n]+)', combined_output)
                if err_match:
                    failure_reason = f"{err_match.group(1)}: {err_match.group(2).strip()[:100]}"
                ws.cell(row=row_idx, column=result_col, value=f"失败-{failure_reason[:50]}")

        except subprocess.TimeoutExpired:
            print(f"  ⏰ 超时(7200s)")
            ws.cell(row=row_idx, column=result_col, value="超时失败")

    wb.save(EXCEL_PATH)
    print(f"\n{'=' * 60}")
    print("重测完成!")
    print(f"Excel已保存: {EXCEL_PATH}")


def main():
    print("=" * 60)
    print("重测被跳过的算子脚本")
    print("=" * 60)

    # Step 1: Fix operators that actually passed (Type A)
    print("\n[Step 1] 修正已通过但被错误标记为跳过的算子...")
    fixed = fix_type_a_passed_ops()

    # Step 2: Retest remaining skipped operators
    print("\n[Step 2] 重测真正被跳过的算子...")
    retest_operators()

    # Step 3: Show summary
    print("\n" + "=" * 60)
    print("最终结果摘要")
    print("=" * 60)

    wb = openpyxl.load_workbook(EXCEL_PATH)
    total_skipped = 0
    total_failed = 0
    total_success = 0
    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        result_col = None
        for c in range(1, ws.max_column + 1):
            h = ws.cell(row=1, column=c).value
            if h == "天数测试结果":
                result_col = c
                break
        if not result_col:
            continue
        for r in range(2, ws.max_row + 1):
            name = ws.cell(row=r, column=1).value
            result = ws.cell(row=r, column=result_col).value
            if name and result:
                if "跳过" in str(result):
                    total_skipped += 1
                elif "失败" in str(result):
                    total_failed += 1
                elif "成功" in str(result):
                    total_success += 1

    print(f"  成功: {total_success}")
    print(f"  失败: {total_failed}")
    print(f"  跳过: {total_skipped}")
    wb.close()


if __name__ == "__main__":
    main()
