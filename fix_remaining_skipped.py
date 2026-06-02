#!/usr/bin/env python3
"""
最终修复剩余的8个跳过算子。

处理逻辑：
1. special_erfc, __ixor__ → 在summary JSON中已通过(acc=True)，直接修正Excel
2. _nested_view_from_buffer_copy → 在summary JSON中skipped，需要重测
3. ConvTranspose, Shape → 从未运行过，需要重测
4. Add_LayerNorm, LayerNorm_GeLU, Linear_GeLU_Linear → 无worktree，保持跳过

步骤1: 从summary JSON中提取已通过的结果
步骤2: 重测有worktree且实际需要跑的算子
"""

import json
import subprocess
import re
import time
import openpyxl
from pathlib import Path
from statistics import mean

EXCEL_PATH = "/root/JudeWorkplace/TianshuOperatorTest/第一批及格算子国产GPU测试.xlsx"
SUMMARY_JSON = "/root/JudeWorkplace/TianshuOperatorTest/test_results_20260509_141455/summary_20260509_141455.json"
FLAGGEMS_BASE = "/root/JudeWorkplace/FlagGems_minimax_2_7"

# Excel算子名 → summary JSON键名映射
EXCEL_TO_SUMMARY_MAP = {
    "aten::special_erfc": "erfc",     # CSV行326, erfc测试 → special_erfc
    "aten::__ixor__": "__xor__",      # CSV行110, __xor__测试 → __ixor__
}


def step1_fix_from_summary():
    """从summary JSON中找出已通过但因命名差异不匹配的算子"""
    print("=" * 60)
    print("步骤1: 通过summary JSON修正命名不匹配的算子")
    print("=" * 60)

    with open(SUMMARY_JSON, "r") as f:
        data = json.load(f)

    wb = openpyxl.load_workbook(EXCEL_PATH)
    fixed = 0

    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        result_col = None
        speed_col = None
        for c in range(1, ws.max_column + 1):
            h = ws.cell(row=1, column=c).value
            if h == "天数测试结果":
                result_col = c
            if h and "生成加速比" in str(h):
                speed_col = c
        if not result_col:
            continue

        for r in range(2, ws.max_row + 1):
            name = ws.cell(row=r, column=1).value
            result = ws.cell(row=r, column=result_col).value
            if not name or not result or "跳过" not in str(result):
                continue

            name_clean = str(name).strip()

            # 检查别名映射
            summary_key = EXCEL_TO_SUMMARY_MAP.get(name_clean)
            if not summary_key:
                continue

            op_data = data["operators"].get(summary_key)
            if op_data and op_data.get("accuracy_passed") == True:
                # 有benchmark数据就更新加速比
                bench = op_data.get("benchmark", {})
                bench_data = bench.get("data") if bench else None
                if bench_data and speed_col:
                    speeds = [d["speedup"] for d in bench_data
                              if d.get("status") == "SUCCESS" and d.get("speedup") is not None]
                    if speeds:
                        avg_speedup = round(mean(speeds), 4)
                        ws.cell(row=r, column=speed_col, value=avg_speedup)
                        print(f"  [{sheet_name}] {name_clean:45s} 加速比={avg_speedup}")

                ws.cell(row=r, column=result_col, value="成功")
                fixed += 1
                print(f"  ✅ [{sheet_name}] {name_clean:45s} → 成功")

    wb.save(EXCEL_PATH)
    print(f"\n步骤1完成: 修正了 {fixed} 个算子")
    return fixed


def find_worktree(op_name):
    """查找算子对应的worktree目录"""
    base = Path(FLAGGEMS_BASE) / ".worktrees"
    candidates = [
        f"gen-{op_name}",
        f"gen-{op_name.replace(' ', '+')}"
    ]
    for c in candidates:
        d = base / c
        if d.is_dir():
            return str(d)
    return None


def get_csv_commands():
    """从CSV读取所有算子命令"""
    import csv
    CSV_PATH = "/root/JudeWorkplace/TianshuOperatorTest/all_operator_commands.csv"
    cmds = {}
    with open(CSV_PATH, "r", encoding="utf-8") as f:
        for row in csv.DictReader(f):
            name = row.get("Operator", "").strip()
            cmds[name] = {
                "test": row.get("Test Command", "").strip(),
                "bench": row.get("Benchmark Command", "").strip(),
            }
    return cmds


def step2_retest():
    """
    步骤2: 重测有worktree且结果缺失的算子
    重点处理: _nested_view_from_buffer_copy (skipped), ConvTranspose(无结果), Shape(无结果)
    """
    print("\n" + "=" * 60)
    print("步骤2: 重测实际被跳过的算子")
    print("=" * 60)

    wb = openpyxl.load_workbook(EXCEL_PATH)
    csv_cmds = get_csv_commands()

    # 需要重测的映射
    retest_plan = {
        # (算子名, sheet名) → CSV中的算子名
        ("aten::_nested_view_from_buffer_copy", "1~5"): "_nested_view_from_buffer_copy",
        ("ConvTranspose", "6"): "ConvTranspose",
        ("Shape", "6"): "Shape",
    }

    # 检查worktree
    available = {}
    for (op_name, _), csv_name in retest_plan.items():
        # 先用CSV名找worktree，不行再用算子名
        wt = find_worktree(csv_name)
        # 如果CSV名不对，试试去掉前缀等
        if not wt:
            wt = find_worktree(op_name.replace("aten::", ""))

        cmd_info = csv_cmds.get(csv_name)

        if wt and cmd_info and cmd_info["test"]:
            available[(op_name, csv_name)] = {
                "worktree": wt,
                "test_cmd": cmd_info["test"],
                "bench_cmd": cmd_info["bench"],
            }
            print(f"  ▶ {op_name:45s} | Worktree: {Path(wt).name}")
        else:
            print(f"  ⏭ {op_name:45s} | {('无worktree' if not wt else '')} {'无CSV命令' if not (cmd_info and cmd_info['test']) else ''}")

    if not available:
        print("  没有可重测的算子")
        wb.close()
        return

    # 开始重测
    print("\n开始执行测试...")
    gpu_id = 0

    for (op_name, csv_name), info in available.items():
        wt_dir = info["worktree"]
        test_cmd = info["test_cmd"]

        # 清理命令中的前缀
        cleaned_cmd = test_cmd
        while True:
            old = cleaned_cmd
            cleaned_cmd = re.sub(r'^(CUDA_VISIBLE_DEVICES=\d+\s+|PYTHONPATH=[^\s]+\s+)+', '', cleaned_cmd)
            if cleaned_cmd == old:
                break

        full_cmd = f"cd {wt_dir} && VLLM_PLUGINS=\"\" CUDA_VISIBLE_DEVICES={gpu_id} {cleaned_cmd}"
        print(f"\n▶ {op_name}")
        print(f"  命令: {cleaned_cmd[:120]}")

        start = time.time()
        try:
            proc = subprocess.run(
                full_cmd, shell=True, capture_output=True, text=True,
                timeout=7200, executable="/bin/bash"
            )
            duration = round(time.time() - start, 1)
            test_passed = (proc.returncode == 0)
            print(f"  {'✅ 通过' if test_passed else '❌ 失败'} ({duration}s, exit={proc.returncode})")

            if not test_passed:
                combined = proc.stdout + "\n" + proc.stderr
                fail_reason = "未知错误"
                m = re.search(r'(RuntimeError|Error|Mismatched elements|TIMEOUT|SIGABRT|Process):\s*([^\n]+)', combined)
                if m:
                    fail_reason = f"{m.group(1)}: {m.group(2).strip()[:120]}"
                print(f"  错误: {fail_reason}")

            # 跑benchmark
            bench_cmd = info["bench_cmd"]
            bench_passed = False
            speedup_val = None
            if bench_cmd and test_passed:
                cleaned_bench = bench_cmd
                while True:
                    old = cleaned_bench
                    cleaned_bench = re.sub(r'^(CUDA_VISIBLE_DEVICES=\d+\s+|PYTHONPATH=[^\s]+\s+)+', '', cleaned_bench)
                    if cleaned_bench == old:
                        break

                full_bench = f"cd {wt_dir} && VLLM_PLUGINS=\"\" CUDA_VISIBLE_DEVICES={gpu_id} {cleaned_bench}"
                print(f"  基准测试: {cleaned_bench[:100]}...")
                bench_proc = subprocess.run(
                    full_bench, shell=True, capture_output=True, text=True,
                    timeout=7200, executable="/bin/bash"
                )
                bench_passed = (bench_proc.returncode == 0)
                print(f"  {'✅ Benchmark通过' if bench_passed else '❌ Benchmark失败'}")

                # 解析加速比
                for line in bench_proc.stdout.split("\n"):
                    m = re.search(r'Gems Speedup.*?(\d+\.\d+)x?', line)
                    if m:
                        speedup_val = float(m.group(1))
                        break
                    m2 = re.search(r'(?:SUCCESS|FAILED)\s+\S+\s+\S+\s+(\d+\.\d+)', line)
                    if m2 and speedup_val is None:
                        speedup_val = float(m2.group(1))

                if speedup_val:
                    print(f"  加速比: {speedup_val}")

            # 更新Excel
            for sheet_name in wb.sheetnames:
                ws = wb[sheet_name]
                result_col = None
                speed_col = None
                for c in range(1, ws.max_column + 1):
                    h = ws.cell(row=1, column=c).value
                    if h == "天数测试结果":
                        result_col = c
                    if h and "生成加速比" in str(h):
                        speed_col = c
                if not result_col:
                    continue

                for r in range(2, ws.max_row + 1):
                    excel_name = ws.cell(row=r, column=1).value
                    if excel_name and str(excel_name).strip() == op_name:
                        if test_passed:
                            ws.cell(row=r, column=result_col, value="成功")
                            if speedup_val and speed_col:
                                ws.cell(row=r, column=speed_col, value=speedup_val)
                        else:
                            ws.cell(row=r, column=result_col, value=f"失败")
                        print(f"  更新Excel[{sheet_name}]: {op_name} → {'成功' if test_passed else '失败'}")
                        break

        except subprocess.TimeoutExpired:
            print(f"  ⏰ 超时(7200s)")
            for sheet_name in wb.sheetnames:
                ws = wb[sheet_name]
                result_col = None
                for c in range(1, ws.max_column + 1):
                    h = ws.cell(row=1, column=c).value
                    if h == "天数测试结果":
                        result_col = c
                        break
                if not result_col:
                    continue
                for r in range(2, ws.max_row + 1):
                    excel_name = ws.cell(row=r, column=1).value
                    if excel_name and str(excel_name).strip() == op_name:
                        ws.cell(row=r, column=result_col, value="超时")
                        print(f"  更新Excel[{sheet_name}]: {op_name} → 超时")
                        break

        # 保存进度
        wb.save(EXCEL_PATH)
        print(f"  Excel已保存")

    wb.close()
    print(f"\n步骤2完成")


def show_summary():
    """显示最终统计"""
    print("\n" + "=" * 60)
    print("最终汇总")
    print("=" * 60)

    wb = openpyxl.load_workbook(EXCEL_PATH, data_only=True)
    total_success = 0
    total_failed = 0
    total_skipped = 0
    total_other = 0

    print("\n剩余跳过算子:")
    print("-" * 60)
    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        result_col = None
        for c in range(1, ws.max_column + 1):
            h = ws.cell(row=1, column=c).value
            if h == "天数测试结果":
                result_col = c
                break
        if not result_col:
            continue
        for r in range(2, ws.max_row + 1):
            name = ws.cell(row=r, column=1).value
            result = ws.cell(row=r, column=result_col).value
            if name and result:
                if "跳过" in str(result):
                    total_skipped += 1
                    if total_skipped <= 5:
                        print(f"  [{sheet_name}] {str(name).strip():40s}")
                elif "失败" in str(result):
                    total_failed += 1
                elif "成功" in str(result):
                    total_success += 1
                else:
                    total_other += 1

    print(f"\n  成功: {total_success}")
    print(f"  失败: {total_failed}")
    print(f"  跳过: {total_skipped}")
    print(f"  其他: {total_other}")
    wb.close()


if __name__ == "__main__":
    step1_fix_from_summary()
    step2_retest()
    show_summary()
