#!/usr/bin/env python3
"""
沐曦 MetaX C550 并行测试脚本
读取 第一批及格算子国产GPU测试.xlsx 中的算子名单，
匹配 all_operator_commands.csv 中的测试命令，
利用 8 块 GPU 并行执行精度测试和 Benchmark，
结果输出到 test_results_YYYYMMDD/ 目录。
"""

import csv
import json
import subprocess
import sys
import os
import re
import time
import threading
from datetime import datetime, timezone
from pathlib import Path
from collections import OrderedDict

# ======================== 配置 ========================
WORK_DIR = Path("/root/JudeWorkplace")
FLAGGEMS_BASE = WORK_DIR / "FlagGems_minimax_2_7"
WORKTREES_DIR = FLAGGEMS_BASE / ".worktrees"
EXCEL_PATH = WORK_DIR / "第一批及格算子国产GPU测试.xlsx"
CSV_PATH = WORK_DIR / "all_operator_commands.csv"

NUM_GPUS = 8
CMD_TIMEOUT = 7200  # 2小时超时

# ======================== 辅助函数 ========================

def log(msg):
    ts = datetime.now().strftime("%H:%M:%S")
    print(f"[{ts}] {msg}", flush=True)


def parse_excel_operators(excel_path):
    """用 openpyxl 解析 Excel 中的算子名"""
    try:
        import openpyxl
    except ImportError:
        log("openpyxl 未安装，尝试 pip install openpyxl ...")
        subprocess.run([sys.executable, "-m", "pip", "install", "openpyxl", "-q"], check=True)
        import openpyxl

    wb = openpyxl.load_workbook(excel_path, data_only=True)
    operators = OrderedDict()

    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        log(f"解析 Sheet: {sheet_name}")
        for row in ws.iter_rows(values_only=True):
            if not row or not row[0]:
                continue
            name = str(row[0]).strip()
            if not name:
                continue
            if name.lower() in ("算子名称", "operator", "算子名称\t生成加速比"):
                continue
            if name.startswith("aten::"):
                name = name[len("aten::"):]
            operators[name] = {"source_sheet": sheet_name}

    log(f"从 Excel 共解析到 {len(operators)} 个算子（去重后）")
    return operators


def parse_csv_commands(csv_path):
    """解析 CSV 中的测试命令"""
    commands = OrderedDict()
    with open(csv_path, "r", encoding="utf-8") as f:
        reader = csv.DictReader(f)
        for row in reader:
            op_name = row.get("Operator", "").strip()
            if not op_name:
                continue
            test_cmd = row.get("Test Command", "").strip()
            bench_cmd = row.get("Benchmark Command", "").strip()
            commands[op_name] = {
                "test_cmd": test_cmd,
                "bench_cmd": bench_cmd,
            }
    log(f"从 CSV 共加载 {len(commands)} 个算子的命令")
    return commands


def get_safe_filename(op_name):
    """将算子名转为安全的文件名"""
    safe = op_name.replace("/", "_").replace("\\", "_").replace(" ", "_")
    safe = safe.replace("<", "_").replace(">", "_").replace(":", "_")
    safe = safe.replace('"', "_").replace("'", "_").replace("|", "_")
    safe = safe.replace("?", "_").replace("*", "_")
    return safe


def strip_gpu_prefix(cmd):
    """去掉命令中已有的 CUDA_VISIBLE_DEVICES= / PYTHONPATH= 等前缀"""
    if not cmd:
        return cmd
    cleaned = re.sub(r'^CUDA_VISIBLE_DEVICES=\d+\s+', '', cmd)
    cleaned = re.sub(r'^PYTHONPATH=[^\s]+\s+', '', cleaned)
    cleaned = re.sub(r'^/usr/bin/python3?\s+', 'python ', cleaned)
    cleaned = re.sub(r'^/opt/conda/bin/python3?\s+', 'python ', cleaned)
    return cleaned.strip()


def find_worktree_dir(op_name):
    """根据算子名找到 worktree 目录"""
    candidates = [
        f"gen-{op_name}",
        f"gen-{op_name.replace(' ', '+')}",
        f"gen-{op_name.replace('+', ' ')}",
        f"gen-{op_name.replace('_', ' ')}",
    ]
    if "+" in op_name:
        candidates.append(f"gen-{op_name}")
    seen = set()
    unique_candidates = []
    for c in candidates:
        if c not in seen:
            seen.add(c)
            unique_candidates.append(c)
    for c in unique_candidates:
        d = WORKTREES_DIR / c
        if d.is_dir():
            return str(d)
    return None


def get_gpu_memory_usage():
    """通过 mx-smi 获取每块 GPU 的显存使用率"""
    try:
        result = subprocess.run(
            ["mx-smi", "--show-memory"],
            capture_output=True, text=True, timeout=30
        )
        output = result.stdout
    except Exception as e:
        log(f"mx-smi 调用失败: {e}")
        return None

    gpu_usages = {}
    current_gpu = None
    for line in output.split("\n"):
        m = re.match(r"GPU#(\d+)\s+", line)
        if m:
            current_gpu = int(m.group(1))
        if current_gpu is not None and "vis_vram usage" in line:
            usage_m = re.search(r"([\d.]+)\s*%", line)
            if usage_m:
                gpu_usages[current_gpu] = float(usage_m.group(1))
                current_gpu = None
    return gpu_usages


def pick_gpu(gpu_lock_dict, gpu_lock):
    """选择当前最空闲的 GPU"""
    with gpu_lock:
        usages = get_gpu_memory_usage()
        if usages is None:
            for i in range(NUM_GPUS):
                if gpu_lock_dict[i] == 0:
                    gpu_lock_dict[i] = 1
                    return i
            return None

        best_gpu = None
        best_usage = 999.0
        for gpu_id in range(NUM_GPUS):
            if gpu_lock_dict[gpu_id] == 0:
                usage = usages.get(gpu_id, 999.0)
                if usage < best_usage:
                    best_usage = usage
                    best_gpu = gpu_id

        if best_gpu is not None:
            gpu_lock_dict[best_gpu] = 1
            return best_gpu
        return None


def release_gpu(gpu_id, gpu_lock_dict, gpu_lock):
    with gpu_lock:
        gpu_lock_dict[gpu_id] = 0


def run_command(cmd, work_dir, gpu_id, output_dir, op_safe_name, cmd_type, timeout=CMD_TIMEOUT):
    """在指定 GPU 上运行命令，同时保存完整输出到日志文件"""
    # 设置 VLLM_PLUGINS="" 避免 vllm 算子注册冲突导致 crash
    full_cmd = f"cd {work_dir} && VLLM_PLUGINS=\"\" CUDA_VISIBLE_DEVICES={gpu_id} {cmd}"
    log(f"  [GPU {gpu_id}] 运行: {cmd[:120]}...")

    # 日志文件路径
    log_file = output_dir / f"{op_safe_name}_{cmd_type}.log"

    start_time = datetime.now(timezone.utc)
    try:
        proc = subprocess.run(
            full_cmd,
            shell=True,
            capture_output=True,
            text=True,
            timeout=timeout,
            executable="/bin/bash",
        )
        end_time = datetime.now(timezone.utc)
        duration = (end_time - start_time).total_seconds()

        # 保存完整日志
        with open(log_file, "w", encoding="utf-8") as f:
            f.write(f"# Command: {full_cmd}\n")
            f.write(f"# Start: {start_time.isoformat()}\n")
            f.write(f"# End: {end_time.isoformat()}\n")
            f.write(f"# Duration: {duration}s\n")
            f.write(f"# Exit Code: {proc.returncode}\n")
            f.write(f"# GPU: {gpu_id}\n")
            f.write("#" + "=" * 60 + "\n\n")
            f.write("=== STDOUT ===\n")
            f.write(proc.stdout)
            f.write("\n\n=== STDERR ===\n")
            f.write(proc.stderr)

        log(f"  [GPU {gpu_id}] 日志已保存: {log_file}")

        stdout_lines = proc.stdout.split("\n")
        stderr_lines = proc.stderr.split("\n")

        tail_lines = 80
        stdout_tail = "\n".join(stdout_lines[-tail_lines:]) if len(stdout_lines) > tail_lines else proc.stdout
        stderr_tail = "\n".join(stderr_lines[-tail_lines:]) if len(stderr_lines) > tail_lines else proc.stderr

        result = {
            "exit_code": proc.returncode,
            "log_file": str(log_file),
            "stdout_tail": stdout_tail,
            "stderr_tail": stderr_tail,
            "start_time": start_time.isoformat(),
            "end_time": end_time.isoformat(),
            "duration_seconds": duration,
            "timed_out": False,
        }

        benchmark_data = parse_benchmark_output(proc.stdout)
        return result, benchmark_data

    except subprocess.TimeoutExpired:
        end_time = datetime.now(timezone.utc)
        duration = (end_time - start_time).total_seconds()
        log(f"  [GPU {gpu_id}] ⏰ 超时 ({timeout}s)")

        # 保存超时日志
        with open(log_file, "w", encoding="utf-8") as f:
            f.write(f"# Command: {full_cmd}\n")
            f.write(f"# Start: {start_time.isoformat()}\n")
            f.write(f"# End: {end_time.isoformat()}\n")
            f.write(f"# Duration: {duration}s\n")
            f.write(f"# Status: TIMEOUT (>{timeout}s)\n")
            f.write("#" + "=" * 60 + "\n")

        return {
            "exit_code": -1,
            "log_file": str(log_file),
            "stdout_tail": "",
            "stderr_tail": "TIMEOUT",
            "start_time": start_time.isoformat(),
            "end_time": end_time.isoformat(),
            "duration_seconds": duration,
            "timed_out": True,
        }, None


def parse_benchmark_output(stdout):
    """
    从 Benchmark 输出中解析 latency/speedup 数据。
    """
    data = []
    current_dtype = None
    in_table = False

    for line in stdout.split("\n"):
        m = re.search(r'dtype=([\w.]+)', line)
        if m:
            current_dtype = m.group(1)

        if "Torch Latency" in line and "Gems Speedup" in line:
            in_table = True
            continue

        if in_table:
            if not line.strip():
                in_table = False
                continue
            if line.startswith("---"):
                continue

            parts = re.split(r'\s{2,}', line.strip())
            if len(parts) >= 5 and parts[0] in ("SUCCESS", "FAILED", "N/A"):
                entry = {
                    "dtype": current_dtype,
                }
                try:
                    entry["status"] = parts[0]
                    entry["torch_latency_ms"] = float(parts[1]) if parts[1] != "N/A" else None
                    entry["gems_latency_ms"] = float(parts[2]) if parts[2] != "N/A" else None
                    entry["speedup"] = float(parts[3]) if parts[3] != "N/A" else None
                    entry["shape"] = parts[4].strip() if len(parts) > 4 else ""
                    entry["size_detail"] = parts[4].strip() if len(parts) > 4 else ""
                except (ValueError, IndexError):
                    entry["raw"] = line.strip()
                data.append(entry)

    return data if data else None


def process_operator(op_name, op_info, csv_commands, results, gpu_lock_dict, gpu_lock, output_dir):
    """处理单个算子的测试"""
    log(f"▶ 处理算子: {op_name} (来源: {op_info.get('source_sheet', '?')})")

    op_safe_name = get_safe_filename(op_name)

    # 1. 查找 CSV 命令
    cmd_info = csv_commands.get(op_name)
    if not cmd_info:
        log(f"  ⚠ CSV 中找不到 {op_name}，跳过")
        results[op_name] = {
            "status": "skipped",
            "skip_reason": "CSV 中未找到匹配命令",
            "source_sheet": op_info.get("source_sheet", ""),
        }
        return

    test_cmd = strip_gpu_prefix(cmd_info.get("test_cmd", ""))
    bench_cmd = strip_gpu_prefix(cmd_info.get("bench_cmd", ""))

    if not test_cmd and not bench_cmd:
        log(f"  ⚠ {op_name} 的 CSV 命令为空，跳过")
        results[op_name] = {
            "status": "skipped",
            "skip_reason": "CSV 命令为空",
            "source_sheet": op_info.get("source_sheet", ""),
        }
        return

    # 2. 查找 worktree 目录
    worktree_path = find_worktree_dir(op_name)
    if not worktree_path:
        log(f"  ⚠ {op_name} 的 worktree 目录不存在，跳过")
        results[op_name] = {
            "status": "skipped",
            "skip_reason": f"Worktree 目录不存在（已搜索: gen-{op_name}）",
            "source_sheet": op_info.get("source_sheet", ""),
        }
        return

    # 3. 分配 GPU
    gpu_id = None
    for retry in range(5):
        gpu_id = pick_gpu(gpu_lock_dict, gpu_lock)
        if gpu_id is not None:
            break
        log(f"  ⏳ GPU 全忙 (retry {retry+1})，等待 30 秒...")
        time.sleep(30)

    if gpu_id is None:
        log(f"  ⚠ {op_name} 无可用 GPU，跳过")
        results[op_name] = {
            "status": "skipped",
            "skip_reason": "无可用 GPU",
            "source_sheet": op_info.get("source_sheet", ""),
        }
        return

    op_start = time.time()

    try:
        log(f"  [GPU {gpu_id}] 开始测试 {op_name}")

        # 4. 运行精度测试
        test_passed = None
        test_info = None
        if test_cmd:
            log(f"  [GPU {gpu_id}] 精度测试: {test_cmd[:100]}...")
            test_result_data, _ = run_command(test_cmd, worktree_path, gpu_id, output_dir, op_safe_name, "test", CMD_TIMEOUT)
            test_passed = test_result_data["exit_code"] == 0 and not test_result_data["timed_out"]
            test_info = {
                "command": test_cmd,
                "passed": test_passed,
                "exit_code": test_result_data["exit_code"],
                "log_file": test_result_data["log_file"],
                "start_time": test_result_data["start_time"],
                "end_time": test_result_data["end_time"],
                "duration_seconds": test_result_data["duration_seconds"],
                "timed_out": test_result_data["timed_out"],
                "stderr_tail": test_result_data["stderr_tail"][-2000:] if test_result_data["stderr_tail"] else "",
            }
        else:
            test_info = {"command": None, "passed": None, "note": "CSV 中无测试命令"}

        # 5. 运行 Benchmark
        bench_passed = None
        bench_info = None
        if bench_cmd:
            log(f"  [GPU {gpu_id}] Benchmark: {bench_cmd[:100]}...")
            bench_result_data, bench_data = run_command(bench_cmd, worktree_path, gpu_id, output_dir, op_safe_name, "bench", CMD_TIMEOUT)
            bench_passed = bench_result_data["exit_code"] == 0 and not bench_result_data["timed_out"]
            bench_info = {
                "command": bench_cmd,
                "passed": bench_passed,
                "exit_code": bench_result_data["exit_code"],
                "log_file": bench_result_data["log_file"],
                "start_time": bench_result_data["start_time"],
                "end_time": bench_result_data["end_time"],
                "duration_seconds": bench_result_data["duration_seconds"],
                "timed_out": bench_result_data["timed_out"],
                "stderr_tail": bench_result_data["stderr_tail"][-2000:] if bench_result_data["stderr_tail"] else "",
                "data": bench_data,
            }
        else:
            bench_info = {"command": None, "passed": None, "note": "CSV 中无 Benchmark 命令"}

        # 6. 汇总状态
        if test_passed is False or bench_passed is False:
            overall_status = "failed"
        elif test_passed is True or bench_passed is True:
            overall_status = "success"
        else:
            overall_status = "skipped"

        test_result = {
            "status": overall_status,
            "gpu_id": gpu_id,
            "worktree_path": worktree_path,
            "source_sheet": op_info.get("source_sheet", ""),
            "duration_seconds": round(time.time() - op_start, 1),
            "accuracy_passed": test_passed,
            "benchmark_passed": bench_passed,
            "test": test_info,
            "benchmark": bench_info,
        }

    except Exception as e:
        log(f"  [GPU {gpu_id}] ❌ {op_name} 异常: {e}")
        import traceback
        traceback.print_exc()
        test_result = {
            "status": "error",
            "gpu_id": gpu_id,
            "worktree_path": worktree_path,
            "source_sheet": op_info.get("source_sheet", ""),
            "error_message": str(e),
            "duration_seconds": round(time.time() - op_start, 1),
        }

    finally:
        release_gpu(gpu_id, gpu_lock_dict, gpu_lock)

    results[op_name] = test_result
    summary_status = test_result['status']
    dur = test_result['duration_seconds']
    log(f"  {'✅' if summary_status == 'success' else '❌' if summary_status == 'failed' else '⚠️'} {op_name} -> {summary_status} (GPU {gpu_id}, {dur:.0f}s)")


def worker_thread(operator_list, csv_commands, shared_results, gpu_lock_dict, gpu_lock, output_dir):
    """工作线程"""
    for op_name, op_info in operator_list:
        if op_name in shared_results:
            continue
        process_operator(op_name, op_info, csv_commands, shared_results, gpu_lock_dict, gpu_lock, output_dir)


def main():
    global_start = datetime.now(timezone.utc)
    log("=" * 60)
    log("沐曦 MetaX C550 并行测试脚本启动")
    log(f"时间戳: {global_start.isoformat()}")
    log("=" * 60)

    # 创建输出目录
    date_str = global_start.strftime("%Y%m%d")
    output_dir = WORK_DIR / f"test_results_{date_str}"
    output_dir.mkdir(parents=True, exist_ok=True)
    log(f"输出目录: {output_dir}")

    # 1. 解析 Excel 算子名单
    log("\n[Step 1/4] 解析 Excel 算子名单...")
    operators = parse_excel_operators(EXCEL_PATH)

    # 2. 解析 CSV 命令
    log("\n[Step 2/4] 解析 CSV 测试命令...")
    csv_commands = parse_csv_commands(CSV_PATH)

    # 3. 匹配并分配
    log("\n[Step 3/4] 匹配算子与命令...")
    matched = []
    unmatched = []
    no_dir = []
    for op_name, op_info in operators.items():
        if op_name not in csv_commands:
            unmatched.append(op_name)
            continue
        worktree_path = find_worktree_dir(op_name)
        if not worktree_path:
            no_dir.append(op_name)
            continue
        matched.append((op_name, op_info))

    log(f"  可执行: {len(matched)} 个")
    log(f"  无 CSV 命令: {len(unmatched)} 个")
    log(f"  无 worktree 目录: {len(no_dir)} 个")

    if len(matched) == 0:
        log("没有可测试的算子，退出")
        return

    log(f"\n  最终测试队列: {len(matched)} 个算子")
    for op_name, _ in matched:
        log(f"    - {op_name}")

    # 4. 并行执行
    log(f"\n[Step 4/4] 使用 {NUM_GPUS} 块 GPU 并行测试...")

    gpu_lock_dict = {i: 0 for i in range(NUM_GPUS)}
    gpu_lock = threading.Lock()
    shared_results = {}

    num_threads = min(NUM_GPUS, len(matched))
    chunk_size = (len(matched) + num_threads - 1) // num_threads
    chunks = [matched[i:i+chunk_size] for i in range(0, len(matched), chunk_size)]

    threads = []
    for chunk in chunks:
        t = threading.Thread(
            target=worker_thread,
            args=(chunk, csv_commands, shared_results, gpu_lock_dict, gpu_lock, output_dir),
        )
        t.start()
        threads.append(t)

    for t in threads:
        t.join()

    global_end = datetime.now(timezone.utc)

    # 5. 汇总结果
    total = len(operators)
    success = sum(1 for v in shared_results.values() if v.get("status") == "success")
    failed = sum(1 for v in shared_results.values() if v.get("status") in ("failed", "error", "partial"))
    skipped = sum(1 for v in shared_results.values() if v.get("status") == "skipped")

    summary = {
        "total_operators_in_excel": total,
        "matched_and_attempted": len(matched),
        "success": success,
        "failed": failed,
        "skipped": skipped,
        "unmatched_no_csv_command": len(unmatched),
        "no_worktree_directory": len(no_dir),
        "total_duration_seconds": round((global_end - global_start).total_seconds(), 1),
    }

    for op_name in unmatched:
        shared_results[op_name] = {
            "status": "skipped",
            "skip_reason": "CSV 中未找到匹配命令",
            "source_sheet": operators[op_name].get("source_sheet", ""),
        }
    for op_name in no_dir:
        shared_results[op_name] = {
            "status": "skipped",
            "skip_reason": "Worktree 目录不存在",
            "source_sheet": operators[op_name].get("source_sheet", ""),
        }

    sorted_results = OrderedDict(sorted(shared_results.items()))

    full_report = {
        "start_time": global_start.isoformat(),
        "end_time": global_end.isoformat(),
        "summary": summary,
        "operators": sorted_results,
    }

    # 6. 保存汇总 JSON
    timestamp = global_start.strftime("%Y%m%d_%H%M%S")
    json_path = output_dir / f"summary_{timestamp}.json"
    with open(json_path, "w", encoding="utf-8") as f:
        json.dump(full_report, f, indent=2, ensure_ascii=False)

    log("\n" + "=" * 60)
    log("测试完成！")
    log(f"结果目录: {output_dir}")
    log(f"汇总文件: {json_path}")
    log(f"Excel 总计: {total} | 成功: {success} | 失败: {failed} | 跳过: {skipped}")
    log(f"总耗时: {summary['total_duration_seconds']:.0f} 秒")
    log("=" * 60)

    # 打印详细结果
    log("\n详细结果:")
    for op_name in sorted(shared_results.keys()):
        r = shared_results[op_name]
        status = r.get("status", "?")
        reason = r.get("skip_reason", "")
        acc = r.get("accuracy_passed")
        bench = r.get("benchmark_passed")
        dur = r.get("duration_seconds")
        if status == "success":
            if acc is True and bench is True:
                log(f"  ✅ {op_name}: 精度通过 + 性能通过 ({dur}s)")
            elif acc is True:
                log(f"  ✅ {op_name}: 精度通过 (无Benchmark) ({dur}s)")
            elif bench is True:
                log(f"  ✅ {op_name}: 性能通过 (无精度测试) ({dur}s)")
        elif status in ("failed", "error", "partial"):
            log(f"  ❌ {op_name}: {status} ({dur}s)")
        else:
            log(f"  ⚠️  {op_name}: 跳过 ({reason})")


if __name__ == "__main__":
    main()