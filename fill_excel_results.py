#!/usr/bin/env python3
"""
读取 summary JSON 文件，统计芯片测试结果，
将平均加速比/失败/跳过等结果写入新的 Excel 文件。

用法: python3 fill_excel_results.py <summary_json_path> [excel_path] [output_excel_path]

示例: python3 fill_excel_results.py /path/to/summary_20260509_141455.json
"""

import json
import sys
import os
from statistics import mean
from pathlib import Path


VENDOR_RESULT_HEADERS = {
    "kunlunxin": "昆仑芯测试结果",
    "tianshu": "天数测试结果",
    "metax": "沐曦测试结果",
    "enflame": "燧原测试结果",
}


def get_result_header(data):
    summary = data.get("summary") or {}
    vendor = str(summary.get("vendor", "")).strip().lower()
    return VENDOR_RESULT_HEADERS.get(vendor, "测试结果")


def build_default_output_path(summary_json_path, excel_path, result_header):
    summary_path = Path(summary_json_path)
    excel_path = Path(excel_path)
    timestamp = summary_path.stem.replace("summary_", "")
    return summary_path.parent / f"{excel_path.stem}_{result_header}_{timestamp}{excel_path.suffix}"


def find_or_create_result_column(ws, header_row, gen_speed_col, result_header):
    for col in range(1, ws.max_column + 1):
        if ws.cell(row=header_row, column=col).value == result_header:
            return col

    new_col = gen_speed_col + 1
    while new_col <= ws.max_column:
        value = ws.cell(row=header_row, column=new_col).value
        if value and "测试结果" in str(value):
            new_col += 1
            continue
        break

    ws.insert_cols(new_col)
    ws.cell(row=header_row, column=new_col, value=result_header)
    return new_col


def fill_excel(summary_json_path, excel_path, output_excel_path=None):
    """读取 summary JSON，分析结果并写入 Excel"""

    # 1. 加载 JSON 数据
    with open(summary_json_path, "r") as f:
        data = json.load(f)

    operators = data.get("operators", {})
    result_header = get_result_header(data)
    print(f"从 JSON 加载了 {len(operators)} 个算子的结果")
    print(f"结果列: {result_header}")

    if output_excel_path is None:
        output_excel_path = build_default_output_path(summary_json_path, excel_path, result_header)

    # 2. 构建算子结果查找表: (sheet_name, op_name) -> result
    op_map = {}
    for op_name, op_info in operators.items():
        sheet = op_info.get("source_sheet", "")
        status = op_info.get("status", "")

        if status == "success":
            # 尝试获取平均加速比
            avg = None
            bench = op_info.get("benchmark")
            if bench and isinstance(bench, dict):
                bench_data = bench.get("data")
                if bench_data and isinstance(bench_data, list):
                    speeds = [
                        d["speedup"]
                        for d in bench_data
                        if d.get("status") == "SUCCESS" and "speedup" in d
                    ]
                    if speeds:
                        avg = mean(speeds)
            if avg is not None:
                result = round(avg, 4)
            else:
                result = "跳过"
        elif status == "failed":
            result = "失败"
        elif status == "skipped":
            result = "跳过"
        else:
            result = status  # error 等

        key = (sheet, op_name)
        op_map[key] = result

    # 3. 打开 Excel
    try:
        import openpyxl
    except ImportError:
        print("openpyxl 未安装，尝试安装...")
        import subprocess
        import sys as _sys
        subprocess.run([_sys.executable, "-m", "pip", "install", "openpyxl", "-q"], check=True)
        import openpyxl

    wb = openpyxl.load_workbook(excel_path)

    # 4. 处理每个 sheet
    sheet_names = ["1~5", "6", "7"]
    total_matched = 0
    total_sheets = 0

    for sheet_name in sheet_names:
        if sheet_name not in wb.sheetnames:
            print(f"  Sheet '{sheet_name}' 不存在于 Excel 中，跳过")
            continue
        ws = wb[sheet_name]
        total_sheets += 1
        print(f"\n=== Sheet: {sheet_name} ===")

        # 找到"生成加速比"列
        gen_speed_col = None
        header_row = None
        for row in range(1, min(ws.max_row + 1, 10)):
            for col in range(1, ws.max_column + 1):
                val = ws.cell(row=row, column=col).value
                if val and "生成加速比" in str(val):
                    gen_speed_col = col
                    header_row = row
                    break
            if gen_speed_col:
                break

        if not gen_speed_col:
            print(f"  找不到 '生成加速比' 列，跳过")
            continue

        # 新增/复用芯片结果列
        new_col = find_or_create_result_column(ws, header_row, gen_speed_col, result_header)
        # 算子名列通常是"生成加速比"列左侧一列
        op_name_col = gen_speed_col - 1

        sheet_matched = 0
        unmatched_ops = []

        for row in range(header_row + 1, ws.max_row + 1):
            op_name = ws.cell(row=row, column=op_name_col).value
            if op_name is None or str(op_name).strip() == "":
                continue
            op_name = str(op_name).strip()
            # 跳过注释行
            if op_name.startswith("（") or op_name.startswith("算子"):
                continue

            key = (sheet_name, op_name)
            if key in op_map:
                result = op_map[key]
                sheet_matched += 1
            else:
                # 尝试带/不带 "aten::" 前缀匹配
                alt_key = None
                if op_name.startswith("aten::"):
                    alt_key = (sheet_name, op_name[6:])
                elif not op_name.startswith("aten::"):
                    alt_key = (sheet_name, "aten::" + op_name)
                if alt_key and alt_key in op_map:
                    result = op_map[alt_key]
                    sheet_matched += 1
                else:
                    result = "未匹配"
                    unmatched_ops.append(op_name)

            ws.cell(
                row=row, column=new_col,
                value=str(result) if isinstance(result, str) else float(result)
            )
            print(f"  {op_name} -> {result}")

        total_matched += sheet_matched
        print(f"  本 sheet 匹配: {sheet_matched}, 未匹配: {len(unmatched_ops)}")
        if unmatched_ops:
            print(f"  未匹配算子: {unmatched_ops}")

    # 5. 保存
    output_excel_path = Path(output_excel_path)
    output_excel_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(output_excel_path)
    print(f"\n完成！共 {total_sheets} 个 sheet，匹配 {total_matched} 个算子")
    print(f"源 Excel: {excel_path}")
    print(f"结果 Excel 已保存: {output_excel_path}")


def main():
    if len(sys.argv) < 2:
        print("用法: python3 fill_excel_results.py <summary_json_path> [excel_path] [output_excel_path]")
        sys.exit(1)

    summary_json_path = sys.argv[1]

    if len(sys.argv) >= 3:
        excel_path = sys.argv[2]
    else:
        # 默认路径
        excel_path = str(Path(__file__).parent / "第一批及格算子国产GPU测试.xlsx")

    output_excel_path = sys.argv[3] if len(sys.argv) >= 4 else None

    if not os.path.exists(summary_json_path):
        print(f"错误: summary JSON 文件不存在: {summary_json_path}")
        sys.exit(1)

    if not os.path.exists(excel_path):
        print(f"错误: Excel 文件不存在: {excel_path}")
        sys.exit(1)

    fill_excel(summary_json_path, excel_path, output_excel_path)


if __name__ == "__main__":
    main()
